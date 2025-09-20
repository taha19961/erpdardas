from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, make_response, g, send_file
from config import Config
from models import db, Car, Employee, Document, CarFile, EmployeeFile, DocumentFile, User, AuditLog, CompanySettings, MaintenanceRecord, Equipment, FuelRecord, EquipmentMaintenance, SalarySettings, EmployeeSalary, AttendanceRecord, OvertimeRecord, AdvancePayment, PayrollRecord
from utils import save_file, log_activity, allowed_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_bcrypt import Bcrypt
from datetime import datetime, date, timedelta
import pandas as pd
import pdfkit
import zipfile
import shutil
import io
import os
from apscheduler.schedulers.background import BackgroundScheduler
from werkzeug.utils import secure_filename
from openpyxl.styles import Font, Alignment

app = Flask(__name__)
app.config.from_object(Config)
db.init_app(app)
bcrypt = Bcrypt(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = "يجب تسجيل الدخول للوصول لهذه الصفحة."
login_manager.login_message_category = "warning"

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.before_request
def create_tables_once():
    if not hasattr(g, 'tables_created'):
        db.create_all()
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'cars'), exist_ok=True)
        os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'employees'), exist_ok=True)
        os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'documents'), exist_ok=True)
        os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'logos'), exist_ok=True)
        if not CompanySettings.query.first():
            settings = CompanySettings()
            db.session.add(settings)
            db.session.commit()
        g.tables_created = True

# الصفحة الرئيسية - لوحة التحكم
@app.route('/')
@login_required
def index():
    # مؤشرات الأداء
    total_equipment_count = Equipment.query.count()
    active_equipment_count = Equipment.query.filter_by(status='active').count()
    # إنتاج اليوم (ستتم إضافته في المرحلة الثانية)
    today_production = 0
    # معدات تحتاج صيانة
    maintenance_alerts = Equipment.query.filter(
        Equipment.next_maintenance_km != None,
        Equipment.current_km >= Equipment.next_maintenance_km
    ).all()
    maintenance_alerts_count = len(maintenance_alerts)
    # مبيعات اليوم (ستتم إضافته في المرحلة الثانية)
    today_sales = 0
    settings = CompanySettings.query.first()
    current_time = datetime.now()
    return render_template('index.html', 
                         total_equipment_count=total_equipment_count,
                         active_equipment_count=active_equipment_count,
                         today_production=today_production,
                         maintenance_alerts=maintenance_alerts,
                         maintenance_alerts_count=maintenance_alerts_count,
                         today_sales=today_sales,
                         settings=settings,
                         current_time=current_time)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user, remember=request.form.get('remember', False))
            log_activity(user, 'login', 'User', user.id, "تسجيل دخول")
            flash('تم تسجيل الدخول بنجاح.', 'success')
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('index'))
        else:
            flash('اسم المستخدم أو كلمة المرور غير صحيحة.', 'danger')
    return render_template('auth/login.html')

@app.route('/logout')
@login_required
def logout():
    log_activity(current_user, 'logout', 'User', current_user.id, "تسجيل خروج")
    logout_user()
    flash('تم تسجيل الخروج.', 'info')
    return redirect(url_for('login'))

@app.route('/init-admin')
def init_admin():
    if User.query.count() == 0:
        admin = User(username='admin', email='admin@company.com')
        admin.set_password('admin123')
        admin.role = 'admin'
        db.session.add(admin)
        db.session.commit()
        return "تم إنشاء المستخدم المسؤول: admin / admin123"
    else:
        return "يوجد مستخدمون مسبقاً."

def get_document_folders():
    base_path = os.path.join(app.config['UPLOAD_FOLDER'], 'documents')
    folders = ['عام']
    if os.path.exists(base_path):
        for item in os.listdir(base_path):
            if os.path.isdir(os.path.join(base_path, item)):
                folders.append(item)
    return sorted(set(folders))

# --- إدارة السيارات ---
@app.route('/cars')
@login_required
def car_list():
    query = request.args.get('q', '').strip()
    if query:
        cars = Car.query.filter(
            db.or_(
                Car.chassis_number.ilike(f'%{query}%'),
                Car.brand.ilike(f'%{query}%'),
                Car.model.ilike(f'%{query}%'),
                Car.plate_number.ilike(f'%{query}%'),
                Car.color.ilike(f'%{query}%'),
                Car.notes.ilike(f'%{query}%')
            )
        ).all()
    else:
        cars = Car.query.all()
    settings = CompanySettings.query.first()
    return render_template('car/list.html', cars=cars, search_query=query, settings=settings)

@app.route('/cars/add', methods=['GET', 'POST'])
@login_required
def add_car():
    if not current_user.can_edit():
        flash('ليس لديك صلاحية لإضافة سيارات.', 'danger')
        return redirect(url_for('car_list'))
    if request.method == 'POST':
        chassis = request.form['chassis_number']
        brand = request.form['brand']
        model = request.form['model']
        car_type = request.form.get('car_type')
        color = request.form.get('color')
        year = request.form.get('year')
        plate = request.form.get('plate_number')
        status = request.form.get('status', 'active')
        notes = request.form.get('notes')
        car = Car(
            chassis_number=chassis,
            brand=brand,
            model=model,
            car_type=car_type,
            color=color,
            year=year,
            plate_number=plate,
            status=status,
            notes=notes
        )
        db.session.add(car)
        db.session.commit()
        files = request.files.getlist('files')
        for file in files:
            filename, filepath = save_file(file, os.path.join(app.config['UPLOAD_FOLDER'], 'cars'), car.unique_id)
            if filename:
                file_type = 'image' if filename.lower().endswith(('png','jpg','jpeg')) else 'pdf'
                car_file = CarFile(filename=filename, filepath=filepath, file_type=file_type, car_id=car.id)
                db.session.add(car_file)
        db.session.commit()
        log_activity(current_user, 'create', 'Car', car.id, f"أضاف سيارة: {brand} {model}")
        flash('تم إضافة السيارة بنجاح!')
        return redirect(url_for('car_list'))
    settings = CompanySettings.query.first()
    return render_template('car/add.html', settings=settings)

@app.route('/cars/<int:car_id>')
@login_required
def car_detail(car_id):
    car = Car.query.get_or_404(car_id)
    settings = CompanySettings.query.first()
    return render_template('car/detail.html', car=car, settings=settings)

@app.route('/cars/<int:car_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_car(car_id):
    if not current_user.can_edit():
        flash('ليس لديك صلاحية تعديل السيارات.', 'danger')
        return redirect(url_for('car_list'))
    car = Car.query.get_or_404(car_id)
    settings = CompanySettings.query.first()
    if request.method == 'POST':
        car.chassis_number = request.form['chassis_number']
        car.brand = request.form['brand']
        car.model = request.form['model']
        car.car_type = request.form.get('car_type')
        car.color = request.form.get('color')
        car.year = request.form.get('year')
        car.plate_number = request.form.get('plate_number')
        car.status = request.form.get('status', 'active')
        car.notes = request.form.get('notes')
        files = request.files.getlist('files')
        for file in files:
            filename, filepath = save_file(file, os.path.join(app.config['UPLOAD_FOLDER'], 'cars'), car.unique_id)
            if filename:
                file_type = 'image' if filename.lower().endswith(('png','jpg','jpeg')) else 'pdf'
                car_file = CarFile(filename=filename, filepath=filepath, file_type=file_type, car_id=car.id)
                db.session.add(car_file)
        db.session.commit()
        log_activity(current_user, 'update', 'Car', car.id, f"عدل سيارة: {car.brand} {car.model}")
        flash('تم تعديل السيارة بنجاح!', 'success')
        return redirect(url_for('car_detail', car_id=car.id))
    return render_template('car/edit.html', car=car, settings=settings)

@app.route('/cars/<int:car_id>/delete', methods=['POST'])
@login_required
def delete_car(car_id):
    if not current_user.can_edit():
        flash('ليس لديك صلاحية حذف السيارات.', 'danger')
        return redirect(url_for('car_list'))
    car = Car.query.get_or_404(car_id)
    car_title = f"{car.brand} {car.model}"
    for file in car.files:
        if os.path.exists(file.filepath):
            os.remove(file.filepath)
    db.session.delete(car)
    db.session.commit()
    log_activity(current_user, 'delete', 'Car', car_id, f"حذف سيارة: {car_title}")
    flash('تم حذف السيارة بنجاح!', 'success')
    return redirect(url_for('car_list'))

@app.route('/cars/<int:car_id>/pdf')
@login_required
def car_pdf(car_id):
    car = Car.query.get_or_404(car_id)
    settings = CompanySettings.query.first()
    html = render_template('car/pdf.html', car=car, settings=settings)
    pdf = pdfkit.from_string(html, False, options={
        'encoding': 'UTF-8',
        'enable-local-file-access': '',
        'quiet': ''
    })
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=car_{car.unique_id}.pdf'
    return response

# --- إدارة سجلات الصيانة للسيارات ---
@app.route('/cars/<int:car_id>/maintenance/add', methods=['GET', 'POST'])
@login_required
def add_maintenance(car_id):
    if not current_user.can_edit():
        flash('ليس لديك صلاحية إضافة سجلات الصيانة.', 'danger')
        return redirect(url_for('car_detail', car_id=car_id))
    car = Car.query.get_or_404(car_id)
    if request.method == 'POST':
        maintenance_type = request.form['maintenance_type']
        date_str = request.form['date']
        cost = request.form.get('cost')
        notes = request.form.get('notes')
        maintenance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        record = MaintenanceRecord(
            car_id=car_id,
            maintenance_type=maintenance_type,
            date=maintenance_date,
            cost=float(cost) if cost else None,
            notes=notes
        )
        db.session.add(record)
        db.session.commit()
        log_activity(current_user, 'create', 'MaintenanceRecord', record.id, f"أضاف صيانة: {maintenance_type} للسيارة {car.brand} {car.model}")
        flash('تم إضافة سجل الصيانة بنجاح!', 'success')
        return redirect(url_for('car_detail', car_id=car_id))
    settings = CompanySettings.query.first()
    return render_template('car/maintenance_add.html', car=car, settings=settings)

@app.route('/cars/<int:car_id>/maintenance/<int:record_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_maintenance(car_id, record_id):
    if not current_user.can_edit():
        flash('ليس لديك صلاحية تعديل سجلات الصيانة.', 'danger')
        return redirect(url_for('car_detail', car_id=car_id))
    car = Car.query.get_or_404(car_id)
    record = MaintenanceRecord.query.get_or_404(record_id)
    if request.method == 'POST':
        record.maintenance_type = request.form['maintenance_type']
        date_str = request.form['date']
        record.date = datetime.strptime(date_str, '%Y-%m-%d').date()
        record.cost = float(request.form['cost']) if request.form.get('cost') else None
        record.notes = request.form.get('notes')
        db.session.commit()
        log_activity(current_user, 'update', 'MaintenanceRecord', record.id, f"عدل صيانة: {record.maintenance_type} للسيارة {car.brand} {car.model}")
        flash('تم تعديل سجل الصيانة بنجاح!', 'success')
        return redirect(url_for('car_detail', car_id=car_id))
    settings = CompanySettings.query.first()
    return render_template('car/maintenance_edit.html', car=car, record=record, settings=settings)

@app.route('/cars/<int:car_id>/maintenance/<int:record_id>/delete', methods=['POST'])
@login_required
def delete_maintenance(car_id, record_id):
    if not current_user.can_edit():
        flash('ليس لديك صلاحية حذف سجلات الصيانة.', 'danger')
        return redirect(url_for('car_detail', car_id=car_id))
    record = MaintenanceRecord.query.get_or_404(record_id)
    maintenance_type = record.maintenance_type
    db.session.delete(record)
    db.session.commit()
    log_activity(current_user, 'delete', 'MaintenanceRecord', record_id, f"حذف صيانة: {maintenance_type}")
    flash('تم حذف سجل الصيانة بنجاح!', 'success')
    return redirect(url_for('car_detail', car_id=car_id))

# --- إدارة الموظفين ---
@app.route('/employees')
@login_required
def employee_list():
    query = request.args.get('q', '').strip()
    if query:
        employees = Employee.query.filter(
            db.or_(
                Employee.full_name.ilike(f'%{query}%'),
                Employee.national_id.ilike(f'%{query}%'),
                Employee.department.ilike(f'%{query}%'),
                Employee.position.ilike(f'%{query}%'),
                Employee.phone.ilike(f'%{query}%'),
                Employee.email.ilike(f'%{query}%'),
                Employee.notes.ilike(f'%{query}%')
            )
        ).all()
    else:
        employees = Employee.query.all()
    settings = CompanySettings.query.first()
    return render_template('employee/list.html', employees=employees, search_query=query, settings=settings)

@app.route('/employees/add', methods=['GET', 'POST'])
@login_required
def add_employee():
    if not current_user.can_edit():
        flash('ليس لديك صلاحية لإضافة موظفين.', 'danger')
        return redirect(url_for('employee_list'))
    if request.method == 'POST':
        national_id = request.form['national_id']
        full_name = request.form['full_name']
        birth_date_str = request.form.get('birth_date')
        gender = request.form.get('gender')
        address = request.form.get('address')
        phone = request.form.get('phone')
        email = request.form.get('email')
        department = request.form.get('department')
        position = request.form.get('position')
        hire_date_str = request.form.get('hire_date')
        status = request.form.get('status', 'active')
        notes = request.form.get('notes')
        birth_date = None
        if birth_date_str:
            birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
        hire_date = None
        if hire_date_str:
            hire_date = datetime.strptime(hire_date_str, '%Y-%m-%d').date()
        emp = Employee(
            national_id=national_id,
            full_name=full_name,
            birth_date=birth_date,
            gender=gender,
            address=address,
            phone=phone,
            email=email,
            department=department,
            position=position,
            hire_date=hire_date,
            status=status,
            notes=notes
        )
        db.session.add(emp)
        db.session.commit()
        files = request.files.getlist('files')
        for file in files:
            filename, filepath = save_file(file, os.path.join(app.config['UPLOAD_FOLDER'], 'employees'), emp.unique_id)
            if filename:
                file_type = 'image' if filename.lower().endswith(('png','jpg','jpeg')) else 'pdf'
                emp_file = EmployeeFile(filename=filename, filepath=filepath, file_type=file_type, employee_id=emp.id)
                db.session.add(emp_file)
        db.session.commit()
        log_activity(current_user, 'create', 'Employee', emp.id, f"أضاف موظف: {full_name}")
        flash('تم إضافة الموظف بنجاح!')
        return redirect(url_for('employee_list'))
    settings = CompanySettings.query.first()
    return render_template('employee/add.html', settings=settings)

@app.route('/employees/<int:employee_id>')
@login_required
def employee_detail(employee_id):
    employee = Employee.query.get_or_404(employee_id)
    settings = CompanySettings.query.first()
    return render_template('employee/detail.html', employee=employee, settings=settings)

@app.route('/employees/<int:employee_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_employee(employee_id):
    if not current_user.can_edit():
        flash('ليس لديك صلاحية تعديل الموظفين.', 'danger')
        return redirect(url_for('employee_list'))
    employee = Employee.query.get_or_404(employee_id)
    settings = CompanySettings.query.first()
    if request.method == 'POST':
        employee.full_name = request.form['full_name']
        employee.national_id = request.form['national_id']
        employee.birth_date = datetime.strptime(request.form['birth_date'], '%Y-%m-%d').date() if request.form.get('birth_date') else None
        employee.gender = request.form.get('gender')
        employee.phone = request.form.get('phone')
        employee.email = request.form.get('email')
        employee.department = request.form.get('department')
        employee.position = request.form.get('position')
        employee.hire_date = datetime.strptime(request.form['hire_date'], '%Y-%m-%d').date() if request.form.get('hire_date') else None
        employee.status = request.form.get('status', 'active')
        employee.notes = request.form.get('notes')
        files = request.files.getlist('files')
        for file in files:
            filename, filepath = save_file(file, os.path.join(app.config['UPLOAD_FOLDER'], 'employees'), employee.unique_id)
            if filename:
                file_type = 'image' if filename.lower().endswith(('png','jpg','jpeg')) else 'pdf'
                emp_file = EmployeeFile(filename=filename, filepath=filepath, file_type=file_type, employee_id=employee.id)
                db.session.add(emp_file)
        db.session.commit()
        log_activity(current_user, 'update', 'Employee', employee.id, f"عدل موظف: {employee.full_name}")
        flash('تم تعديل الموظف بنجاح!', 'success')
        return redirect(url_for('employee_detail', employee_id=employee.id))
    return render_template('employee/edit.html', employee=employee, settings=settings)

@app.route('/employees/<int:employee_id>/delete', methods=['POST'])
@login_required
def delete_employee(employee_id):
    if not current_user.can_edit():
        flash('ليس لديك صلاحية حذف الموظفين.', 'danger')
        return redirect(url_for('employee_list'))
    employee = Employee.query.get_or_404(employee_id)
    emp_name = employee.full_name
    for file in employee.files:
        if os.path.exists(file.filepath):
            os.remove(file.filepath)
    db.session.delete(employee)
    db.session.commit()
    log_activity(current_user, 'delete', 'Employee', employee_id, f"حذف موظف: {emp_name}")
    flash('تم حذف الموظف بنجاح!', 'success')
    return redirect(url_for('employee_list'))

@app.route('/employees/<int:employee_id>/pdf')
@login_required
def employee_pdf(employee_id):
    employee = Employee.query.get_or_404(employee_id)
    settings = CompanySettings.query.first()
    html = render_template('employee/pdf.html', employee=employee, settings=settings)
    pdf = pdfkit.from_string(html, False, options={
        'encoding': 'UTF-8',
        'enable-local-file-access': '',
        'quiet': ''
    })
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=employee_{employee.unique_id}.pdf'
    return response

# --- إدارة الوثائق ---
@app.route('/documents')
@login_required
def document_list():
    folder_filter = request.args.get('folder')
    query = request.args.get('q', '').strip()
    documents = Document.query
    if folder_filter:
        documents = documents.filter(Document.folder == folder_filter)
    if query:
        documents = documents.filter(
            db.or_(
                Document.title.ilike(f'%{query}%'),
                Document.doc_type.ilike(f'%{query}%'),
                Document.source.ilike(f'%{query}%'),
                Document.folder.ilike(f'%{query}%'),
                Document.notes.ilike(f'%{query}%'),
                Document.unique_id.ilike(f'%{query}%')
            )
        )
    documents = documents.all()
    folders = get_document_folders()
    settings = CompanySettings.query.first()
    return render_template('document/list.html', documents=documents, search_query=query, folders=folders, settings=settings)

@app.route('/documents/add', methods=['GET', 'POST'])
@login_required
def add_document():
    if not current_user.can_edit():
        flash('ليس لديك صلاحية لإضافة وثائق.', 'danger')
        return redirect(url_for('document_list'))
    folders = get_document_folders()
    settings = CompanySettings.query.first()
    if request.method == 'POST':
        title = request.form['title']
        doc_type = request.form.get('doc_type')
        source = request.form.get('source')
        issue_date_str = request.form.get('issue_date')
        receive_date_str = request.form.get('receive_date')
        expiry_date_str = request.form.get('expiry_date')
        status = request.form.get('status', 'pending')
        folder = request.form.get('folder', 'عام')
        new_folder = request.form.get('new_folder')
        notes = request.form.get('notes')
        if new_folder:
            folder = new_folder
        issue_date = None
        if issue_date_str:
            issue_date = datetime.strptime(issue_date_str, '%Y-%m-%d').date()
        receive_date = None
        if receive_date_str:
            receive_date = datetime.strptime(receive_date_str, '%Y-%m-%d').date()
        expiry_date = None
        if expiry_date_str:
            expiry_date = datetime.strptime(expiry_date_str, '%Y-%m-%d').date()
        doc = Document(
            title=title,
            doc_type=doc_type,
            source=source,
            issue_date=issue_date,
            receive_date=receive_date,
            expiry_date=expiry_date,
            status=status,
            folder=folder,
            notes=notes
        )
        db.session.add(doc)
        db.session.commit()
        folder_path = os.path.join(app.config['UPLOAD_FOLDER'], 'documents', folder)
        files = request.files.getlist('files')
        for file in files:
            filename, filepath = save_file(file, folder_path, doc.unique_id)
            if filename:
                file_type = 'image' if filename.lower().endswith(('png','jpg','jpeg')) else 'pdf'
                doc_file = DocumentFile(filename=filename, filepath=filepath, file_type=file_type, document_id=doc.id)
                db.session.add(doc_file)
        db.session.commit()
        log_activity(current_user, 'create', 'Document', doc.id, f"أضاف وثيقة: {title}")
        flash('تم إضافة الوثيقة بنجاح!')
        return redirect(url_for('document_list'))
    return render_template('document/add.html', folders=folders, settings=settings)

@app.route('/documents/<int:document_id>')
@login_required
def document_detail(document_id):
    document = Document.query.get_or_404(document_id)
    settings = CompanySettings.query.first()
    return render_template('document/detail.html', document=document, settings=settings)

@app.route('/documents/<int:document_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_document(document_id):
    if not current_user.can_edit():
        flash('ليس لديك صلاحية تعديل الوثائق.', 'danger')
        return redirect(url_for('document_list'))
    document = Document.query.get_or_404(document_id)
    folders = get_document_folders()
    settings = CompanySettings.query.first()
    if request.method == 'POST':
        document.title = request.form['title']
        document.doc_type = request.form.get('doc_type')
        document.source = request.form.get('source')
        document.issue_date = datetime.strptime(request.form['issue_date'], '%Y-%m-%d').date() if request.form.get('issue_date') else None
        document.receive_date = datetime.strptime(request.form['receive_date'], '%Y-%m-%d').date() if request.form.get('receive_date') else None
        document.expiry_date = datetime.strptime(request.form['expiry_date'], '%Y-%m-%d').date() if request.form.get('expiry_date') else None
        document.status = request.form.get('status', 'pending')
        document.folder = request.form.get('folder', 'عام')
        document.notes = request.form.get('notes')
        new_folder = request.form.get('new_folder')
        if new_folder:
            document.folder = new_folder
        folder_path = os.path.join(app.config['UPLOAD_FOLDER'], 'documents', document.folder)
        files = request.files.getlist('files')
        for file in files:
            filename, filepath = save_file(file, folder_path, document.unique_id)
            if filename:
                file_type = 'image' if filename.lower().endswith(('png','jpg','jpeg')) else 'pdf'
                doc_file = DocumentFile(filename=filename, filepath=filepath, file_type=file_type, document_id=document.id)
                db.session.add(doc_file)
        db.session.commit()
        log_activity(current_user, 'update', 'Document', document.id, f"عدل وثيقة: {document.title}")
        flash('تم تعديل الوثيقة بنجاح!', 'success')
        return redirect(url_for('document_detail', document_id=document.id))
    return render_template('document/edit.html', document=document, folders=folders, settings=settings)

@app.route('/documents/<int:document_id>/delete', methods=['POST'])
@login_required
def delete_document(document_id):
    if not current_user.can_edit():
        flash('ليس لديك صلاحية حذف الوثائق.', 'danger')
        return redirect(url_for('document_list'))
    document = Document.query.get_or_404(document_id)
    doc_title = document.title
    for file in document.files:
        if os.path.exists(file.filepath):
            os.remove(file.filepath)
    db.session.delete(document)
    db.session.commit()
    log_activity(current_user, 'delete', 'Document', document_id, f"حذف وثيقة: {doc_title}")
    flash('تم حذف الوثيقة بنجاح!', 'success')
    return redirect(url_for('document_list'))

@app.route('/documents/<int:document_id>/pdf')
@login_required
def document_pdf(document_id):
    document = Document.query.get_or_404(document_id)
    settings = CompanySettings.query.first()
    html = render_template('document/pdf.html', document=document, settings=settings)
    pdf = pdfkit.from_string(html, False, options={
        'encoding': 'UTF-8',
        'enable-local-file-access': '',
        'quiet': ''
    })
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=document_{document.unique_id}.pdf'
    return response

# --- خدمة الملفات ---
@app.route('/uploads/<path:filename>')
@login_required
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# --- الإشعارات ---
@app.route('/notifications')
@login_required
def notifications():
    today = date.today()
    next_week = today + timedelta(days=7)
    expiring_docs = Document.query.filter(
        Document.expiry_date.between(today, next_week),
        Document.status != 'expired'
    ).all()
    expired_docs = Document.query.filter(
        Document.expiry_date < today,
        Document.status != 'expired'
    ).all()
    for doc in expired_docs:
        doc.status = 'expired'
    db.session.commit()
    settings = CompanySettings.query.first()
    return render_template('notifications.html',
                         expiring_docs=expiring_docs,
                         expired_docs=expired_docs,
                         settings=settings)

# --- النسخ الاحتياطي ---
def backup_system():
    backup_dir = os.path.join(os.getcwd(), 'backups')
    os.makedirs(backup_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_filename = f"backup_{timestamp}.zip"
    backup_path = os.path.join(backup_dir, backup_filename)
    try:
        with zipfile.ZipFile(backup_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            db_path = 'archive.db'
            if os.path.exists(db_path):
                zipf.write(db_path, arcname='archive.db')
            uploads_path = app.config['UPLOAD_FOLDER']
            if os.path.exists(uploads_path):
                for root, dirs, files in os.walk(uploads_path):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, os.getcwd())
                        zipf.write(file_path, arcname=arcname)
        print(f"[Backup] تم إنشاء نسخة احتياطية: {backup_filename}")
    except Exception as e:
        print(f"[Backup Error] {str(e)}")

scheduler = BackgroundScheduler()
scheduler.add_job(func=backup_system, trigger="cron", hour=2, minute=0)
scheduler.start()

@app.route('/backups')
@login_required
def backup_list():
    if not current_user.is_admin():
        flash('ليس لديك صلاحية الوصول لهذه الصفحة.', 'danger')
        return redirect(url_for('index'))
    backup_dir = os.path.join(os.getcwd(), 'backups')
    backups = []
    if os.path.exists(backup_dir):
        for filename in os.listdir(backup_dir):
            if filename.endswith('.zip'):
                filepath = os.path.join(backup_dir, filename)
                size = os.path.getsize(filepath)
                created = os.path.getctime(filepath)
                backups.append({
                    'filename': filename,
                    'size': f"{size / (1024*1024):.2f} MB",
                    'created': datetime.fromtimestamp(created).strftime('%Y-%m-%d %H:%M:%S')
                })
    backups.sort(key=lambda x: x['created'], reverse=True)
    settings = CompanySettings.query.first()
    return render_template('backup/list.html', backups=backups, settings=settings)

@app.route('/backups/trigger')
@login_required
def trigger_backup():
    if not current_user.is_admin():
        flash('ليس لديك صلاحية.', 'danger')
        return redirect(url_for('index'))
    backup_system()
    flash('تم إنشاء نسخة احتياطية يدوياً.', 'success')
    return redirect(url_for('backup_list'))

@app.route('/backups/download/<filename>')
@login_required
def download_backup(filename):
    if not current_user.is_admin():
        flash('ليس لديك صلاحية.', 'danger')
        return redirect(url_for('index'))
    return send_from_directory('backups', filename, as_attachment=True)

@app.route('/backups/delete/<filename>')
@login_required
def delete_backup(filename):
    if not current_user.is_admin():
        flash('ليس لديك صلاحية.', 'danger')
        return redirect(url_for('index'))
    filepath = os.path.join('backups', filename)
    if os.path.exists(filepath):
        os.remove(filepath)
        flash('تم حذف النسخة الاحتياطية.', 'success')
    else:
        flash('الملف غير موجود.', 'warning')
    return redirect(url_for('backup_list'))

# --- إدارة المستخدمين ---
@app.route('/users')
@login_required
def user_list():
    if not current_user.is_admin():
        flash('ليس لديك صلاحية عرض المستخدمين.', 'danger')
        return redirect(url_for('index'))
    users = User.query.all()
    settings = CompanySettings.query.first()
    return render_template('user/list.html', users=users, settings=settings)

@app.route('/users/add', methods=['GET', 'POST'])
@login_required
def add_user():
    if not current_user.is_admin():
        flash('ليس لديك صلاحية لإضافة مستخدمين.', 'danger')
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']
        role = request.form['role']
        if User.query.filter_by(username=username).first():
            flash('اسم المستخدم موجود مسبقاً.', 'warning')
            return render_template('user/add.html')
        if User.query.filter_by(email=email).first():
            flash('البريد الإلكتروني موجود مسبقاً.', 'warning')
            return render_template('user/add.html')
        new_user = User(username=username, email=email, role=role)
        new_user.set_password(password)
        db.session.add(new_user)
        db.session.commit()
        log_activity(current_user, 'create', 'User', new_user.id, f"أضاف مستخدم: {username}")
        flash(f'تم إنشاء المستخدم {username} بنجاح!', 'success')
        return redirect(url_for('user_list'))
    settings = CompanySettings.query.first()
    return render_template('user/add.html', settings=settings)

# --- سجل النشاط ---
@app.route('/audit')
@login_required
def audit_log():
    if not current_user.is_admin():
        flash('ليس لديك صلاحية عرض سجل النشاط.', 'danger')
        return redirect(url_for('index'))
    logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).all()
    settings = CompanySettings.query.first()
    return render_template('audit/list.html', logs=logs, settings=settings)

# --- إعدادات الشركة ---
@app.route('/settings/company', methods=['GET', 'POST'])
@login_required
def company_settings():
    if not current_user.is_admin():
        flash('ليس لديك صلاحية الوصول لهذه الصفحة.', 'danger')
        return redirect(url_for('index'))
    settings = CompanySettings.query.first()
    if not settings:
        settings = CompanySettings()
        db.session.add(settings)
        db.session.commit()
    if request.method == 'POST':
        company_name = request.form['company_name']
        settings.company_name = company_name
        if 'logo' in request.files:
            file = request.files['logo']
            if file and allowed_file(file.filename, {'png', 'jpg', 'jpeg'}):
                if settings.logo_filename and settings.logo_filename != "default-logo.png":
                    old_logo_path = os.path.join(app.config['UPLOAD_FOLDER'], 'logos', settings.logo_filename)
                    if os.path.exists(old_logo_path):
                        os.remove(old_logo_path)
                filename = secure_filename(file.filename)
                unique_filename = f"logo_{int(datetime.now().timestamp())}_{filename}"
                logo_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'logos')
                os.makedirs(logo_dir, exist_ok=True)
                filepath = os.path.join(logo_dir, unique_filename)
                file.save(filepath)
                settings.logo_filename = unique_filename
        db.session.commit()
        flash('تم تحديث إعدادات الشركة بنجاح.', 'success')
        return redirect(url_for('company_settings'))
    return render_template('settings/company.html', settings=settings)

# --- تصدير Excel ---
@app.route('/cars/export')
@login_required
def export_cars():
    cars = Car.query.all()
    data = [{
        'الرقم المرجعي': car.unique_id,
        'رقم الشاسيه': car.chassis_number,
        'الماركة': car.brand,
        'الموديل': car.model,
        'النوع': car.car_type,
        'اللون': car.color,
        'السنة': car.year,
        'رقم اللوحة': car.plate_number,
        'الحالة': car.status,
        'تاريخ الإدخال': car.created_at.strftime('%Y-%m-%d'),
        'ملاحظات': car.notes
    } for car in cars]
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='السيارات')
        workbook = writer.book
        worksheet = writer.sheets['السيارات']
        settings = CompanySettings.query.first()
        company_name = settings.company_name if settings else "شركة الأرشيف"
        worksheet['A1'] = company_name
        worksheet['A1'].font = Font(size=16, bold=True, color="0070C0")
        worksheet.merge_cells('A1:L1')
        worksheet['A1'].alignment = Alignment(horizontal="center")
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name='cars_export.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/employees/export')
@login_required
def export_employees():
    employees = Employee.query.all()
    data = [{
        'الرقم المرجعي': emp.unique_id,
        'الاسم الكامل': emp.full_name,
        'الرقم الوطني': emp.national_id,
        'تاريخ الميلاد': emp.birth_date.strftime('%Y-%m-%d') if emp.birth_date else '',
        'الجنس': emp.gender,
        'الهاتف': emp.phone,
        'البريد الإلكتروني': emp.email,
        'القسم': emp.department,
        'الوظيفة': emp.position,
        'تاريخ التعيين': emp.hire_date.strftime('%Y-%m-%d') if emp.hire_date else '',
        'الحالة': emp.status,
        'تاريخ الإدخال': emp.created_at.strftime('%Y-%m-%d'),
        'ملاحظات': emp.notes
    } for emp in employees]
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='الموظفون')
        workbook = writer.book
        worksheet = writer.sheets['الموظفون']
        settings = CompanySettings.query.first()
        company_name = settings.company_name if settings else "شركة الأرشيف"
        worksheet['A1'] = company_name
        worksheet['A1'].font = Font(size=16, bold=True, color="0070C0")
        worksheet.merge_cells('A1:M1')
        worksheet['A1'].alignment = Alignment(horizontal="center")
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name='employees_export.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/documents/export')
@login_required
def export_documents():
    documents = Document.query.all()
    data = [{
        'الرقم المرجعي': doc.unique_id,
        'العنوان': doc.title,
        'النوع': doc.doc_type,
        'الجهة المصدرة': doc.source,
        'تاريخ الإصدار': doc.issue_date.strftime('%Y-%m-%d') if doc.issue_date else '',
        'تاريخ الاستلام': doc.receive_date.strftime('%Y-%m-%d'),
        'المجلد': doc.folder,
        'الحالة': doc.status,
        'تاريخ الإدخال': doc.created_at.strftime('%Y-%m-%d'),
        'ملاحظات': doc.notes
    } for doc in documents]
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='الوثائق')
        workbook = writer.book
        worksheet = writer.sheets['الوثائق']
        settings = CompanySettings.query.first()
        company_name = settings.company_name if settings else "شركة الأرشيف"
        worksheet['A1'] = company_name
        worksheet['A1'].font = Font(size=16, bold=True, color="0070C0")
        worksheet.merge_cells('A1:K1')
        worksheet['A1'].alignment = Alignment(horizontal="center")
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name='documents_export.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# --- إدارة المعدات ---
@app.route('/equipment')
@login_required
def equipment_list():
    query = request.args.get('q', '').strip()
    if query:
        equipment = Equipment.query.filter(
            db.or_(
                Equipment.equipment_type.ilike(f'%{query}%'),
                Equipment.brand.ilike(f'%{query}%'),
                Equipment.model.ilike(f'%{query}%'),
                Equipment.chassis_number.ilike(f'%{query}%'),
                Equipment.status.ilike(f'%{query}%')
            )
        ).all()
    else:
        equipment = Equipment.query.all()
    settings = CompanySettings.query.first()
    return render_template('equipment/list.html', equipment=equipment, search_query=query, settings=settings)

@app.route('/equipment/add', methods=['GET', 'POST'])
@login_required
def add_equipment():
    if not current_user.can_edit():
        flash('ليس لديك صلاحية لإضافة معدات.', 'danger')
        return redirect(url_for('equipment_list'))
    if request.method == 'POST':
        equipment_type = request.form['equipment_type']
        brand = request.form['brand']
        model = request.form['model']
        chassis_number = request.form['chassis_number']
        engine_number = request.form.get('engine_number')
        capacity = request.form.get('capacity')
        max_load = request.form.get('max_load')
        current_km = request.form.get('current_km', 0)
        next_maintenance_km = request.form.get('next_maintenance_km')
        status = request.form.get('status', 'active')
        purchase_date_str = request.form.get('purchase_date')
        notes = request.form.get('notes')
        purchase_date = None
        if purchase_date_str:
            purchase_date = datetime.strptime(purchase_date_str, '%Y-%m-%d').date()
        equipment = Equipment(
            equipment_type=equipment_type,
            brand=brand,
            model=model,
            chassis_number=chassis_number,
            engine_number=engine_number,
            capacity=float(capacity) if capacity else None,
            max_load=float(max_load) if max_load else None,
            current_km=int(current_km) if current_km else 0,
            last_maintenance_km=int(current_km) if current_km else 0,
            next_maintenance_km=int(next_maintenance_km) if next_maintenance_km else None,
            status=status,
            purchase_date=purchase_date,
            notes=notes
        )
        db.session.add(equipment)
        db.session.commit()
        log_activity(current_user, 'create', 'Equipment', equipment.id, f"أضاف معدة: {brand} {model}")
        flash('تم إضافة المعدة بنجاح!', 'success')
        return redirect(url_for('equipment_list'))
    settings = CompanySettings.query.first()
    return render_template('equipment/add.html', settings=settings)

@app.route('/equipment/<int:equipment_id>')
@login_required
def equipment_detail(equipment_id):
    equipment = Equipment.query.get_or_404(equipment_id)
    settings = CompanySettings.query.first()
    return render_template('equipment/detail.html', equipment=equipment, settings=settings)

@app.route('/equipment/<int:equipment_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_equipment(equipment_id):
    if not current_user.can_edit():
        flash('ليس لديك صلاحية تعديل المعدات.', 'danger')
        return redirect(url_for('equipment_list'))
    equipment = Equipment.query.get_or_404(equipment_id)
    settings = CompanySettings.query.first()
    if request.method == 'POST':
        equipment.equipment_type = request.form['equipment_type']
        equipment.brand = request.form['brand']
        equipment.model = request.form['model']
        equipment.chassis_number = request.form['chassis_number']
        equipment.engine_number = request.form.get('engine_number')
        equipment.capacity = float(request.form['capacity']) if request.form.get('capacity') else None
        equipment.max_load = float(request.form['max_load']) if request.form.get('max_load') else None
        equipment.current_km = int(request.form['current_km']) if request.form.get('current_km') else 0
        equipment.next_maintenance_km = int(request.form['next_maintenance_km']) if request.form.get('next_maintenance_km') else None
        equipment.status = request.form.get('status', 'active')
        equipment.purchase_date = datetime.strptime(request.form['purchase_date'], '%Y-%m-%d').date() if request.form.get('purchase_date') else None
        equipment.notes = request.form.get('notes')
        db.session.commit()
        log_activity(current_user, 'update', 'Equipment', equipment.id, f"عدل معدة: {equipment.brand} {equipment.model}")
        flash('تم تعديل المعدة بنجاح!', 'success')
        return redirect(url_for('equipment_detail', equipment_id=equipment.id))
    return render_template('equipment/edit.html', equipment=equipment, settings=settings)

@app.route('/equipment/<int:equipment_id>/delete', methods=['POST'])
@login_required
def delete_equipment(equipment_id):
    if not current_user.can_edit():
        flash('ليس لديك صلاحية حذف المعدات.', 'danger')
        return redirect(url_for('equipment_list'))
    equipment = Equipment.query.get_or_404(equipment_id)
    equipment_name = f"{equipment.brand} {equipment.model}"
    db.session.delete(equipment)
    db.session.commit()
    log_activity(current_user, 'delete', 'Equipment', equipment_id, f"حذف معدة: {equipment_name}")
    flash('تم حذف المعدة بنجاح!', 'success')
    return redirect(url_for('equipment_list'))

# --- إدارة الوقود ---
@app.route('/equipment/<int:equipment_id>/fuel/add', methods=['GET', 'POST'])
@login_required
def add_fuel_record(equipment_id):
    if not current_user.can_edit():
        flash('ليس لديك صلاحية إضافة سجلات الوقود.', 'danger')
        return redirect(url_for('equipment_detail', equipment_id=equipment_id))
    equipment = Equipment.query.get_or_404(equipment_id)
    if request.method == 'POST':
        date_str = request.form['date']
        quantity = request.form['quantity']
        price_per_liter = request.form['price_per_liter']
        total_cost = request.form['total_cost']
        current_km = request.form['current_km']
        fuel_type = request.form.get('fuel_type')
        notes = request.form.get('notes')
        fuel_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        fuel_record = FuelRecord(
            equipment_id=equipment_id,
            date=fuel_date,
            quantity=float(quantity),
            price_per_liter=float(price_per_liter),
            total_cost=float(total_cost),
            current_km=int(current_km),
            fuel_type=fuel_type,
            notes=notes
        )
        equipment.current_km = int(current_km)
        db.session.add(fuel_record)
        db.session.commit()
        log_activity(current_user, 'create', 'FuelRecord', fuel_record.id, f"أضاف وقود للمعدة: {equipment.brand} {equipment.model}")
        flash('تم إضافة سجل الوقود بنجاح!', 'success')
        return redirect(url_for('equipment_detail', equipment_id=equipment_id))
    settings = CompanySettings.query.first()
    return render_template('equipment/fuel_add.html', equipment=equipment, settings=settings)

# --- إدارة صيانة المعدات ---
@app.route('/equipment/<int:equipment_id>/maintenance/add', methods=['GET', 'POST'])
@login_required
def add_equipment_maintenance(equipment_id):
    if not current_user.can_edit():
        flash('ليس لديك صلاحية إضافة سجلات الصيانة.', 'danger')
        return redirect(url_for('equipment_detail', equipment_id=equipment_id))
    equipment = Equipment.query.get_or_404(equipment_id)
    if request.method == 'POST':
        maintenance_type = request.form['maintenance_type']
        description = request.form['description']
        date_str = request.form['date']
        cost = request.form.get('cost')
        current_km = request.form['current_km']
        next_maintenance_km = request.form.get('next_maintenance_km')
        performed_by = request.form.get('performed_by')
        notes = request.form.get('notes')
        maintenance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        maintenance = EquipmentMaintenance(
            equipment_id=equipment_id,
            maintenance_type=maintenance_type,
            description=description,
            date=maintenance_date,
            cost=float(cost) if cost else None,
            current_km=int(current_km),
            next_maintenance_km=int(next_maintenance_km) if next_maintenance_km else None,
            performed_by=performed_by,
            notes=notes
        )
        equipment.last_maintenance_km = int(current_km)
        if next_maintenance_km:
            equipment.next_maintenance_km = int(next_maintenance_km)
        db.session.add(maintenance)
        db.session.commit()
        log_activity(current_user, 'create', 'EquipmentMaintenance', maintenance.id, f"أضاف صيانة للمعدة: {equipment.brand} {equipment.model}")
        flash('تم إضافة سجل الصيانة بنجاح!', 'success')
        return redirect(url_for('equipment_detail', equipment_id=equipment_id))
    settings = CompanySettings.query.first()
    return render_template('equipment/maintenance_add.html', equipment=equipment, settings=settings)

# --- تنبيهات الصيانة ---
@app.route('/equipment/maintenance-alerts')
@login_required
def maintenance_alerts():
    alerts = Equipment.query.filter(
        Equipment.next_maintenance_km != None,
        Equipment.current_km >= Equipment.next_maintenance_km
    ).all()
    settings = CompanySettings.query.first()
    return render_template('equipment/maintenance_alerts.html', alerts=alerts, settings=settings)

# --- إدارة الرواتب ---
# --- إدارة إعدادات الرواتب ---
@app.route('/salary/settings', methods=['GET', 'POST'])
@login_required
def salary_settings():
    if not current_user.is_admin():
        flash('ليس لديك صلاحية الوصول لهذه الصفحة.', 'danger')
        return redirect(url_for('index'))
    settings = SalarySettings.query.first()
    if not settings:
        settings = SalarySettings()
        db.session.add(settings)
        db.session.commit()
    if request.method == 'POST':
        settings.daily_rate = float(request.form['daily_rate'])
        settings.hourly_rate = float(request.form['hourly_rate'])
        settings.overtime_daily_rate = float(request.form['overtime_daily_rate'])
        settings.overtime_hourly_rate = float(request.form['overtime_hourly_rate'])
        db.session.commit()
        flash('تم تحديث إعدادات الرواتب بنجاح!', 'success')
        return redirect(url_for('salary_settings'))
    settings = CompanySettings.query.first()
    return render_template('salary/settings.html', salary_settings=settings, settings=settings)

# --- إدارة رواتب الموظفين ---
@app.route('/salary/employees')
@login_required
def salary_employees_list():
    employees = Employee.query.all()
    settings = CompanySettings.query.first()
    return render_template('salary/employees_list.html', employees=employees, settings=settings)

@app.route('/salary/employee/<int:employee_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_employee_salary(employee_id):
    if not current_user.is_admin():
        flash('ليس لديك صلاحية تعديل رواتب الموظفين.', 'danger')
        return redirect(url_for('salary_employees_list'))
    employee = Employee.query.get_or_404(employee_id)
    salary_info = EmployeeSalary.query.filter_by(employee_id=employee_id).first()
    if not salary_info:
        salary_info = EmployeeSalary(employee_id=employee_id)
        db.session.add(salary_info)
    if request.method == 'POST':
        salary_info.base_salary = float(request.form['base_salary']) if request.form.get('base_salary') else 0
        salary_info.daily_wage = float(request.form['daily_wage']) if request.form.get('daily_wage') else None
        salary_info.hourly_wage = float(request.form['hourly_wage']) if request.form.get('hourly_wage') else None
        salary_info.notes = request.form.get('notes')
        db.session.commit()
        flash('تم تحديث معلومات الراتب بنجاح!', 'success')
        return redirect(url_for('salary_employees_list'))
    settings = CompanySettings.query.first()
    return render_template('salary/employee_salary_edit.html', employee=employee, salary_info=salary_info, settings=settings)

# --- تسجيل الحضور والغياب الجماعي ---
@app.route('/salary/attendance/bulk', methods=['GET', 'POST'])
@login_required
def bulk_attendance():
    if not current_user.can_edit():
        flash('ليس لديك صلاحية تسجيل الحضور والغياب.', 'danger')
        return redirect(url_for('index'))
    employees = Employee.query.filter_by(status='active').all()
    if request.method == 'POST':
        date_str = request.form['date']
        attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        week_number = attendance_date.isocalendar()[1]
        year = attendance_date.year
        for employee in employees:
            status = request.form.get(f'status_{employee.id}')
            notes = request.form.get(f'notes_{employee.id}')
            if status:  # فقط إذا تم اختيار حالة
                existing = AttendanceRecord.query.filter_by(employee_id=employee.id, date=attendance_date).first()
                if existing:
                    existing.status = status
                    existing.notes = notes
                else:
                    attendance = AttendanceRecord(
                        employee_id=employee.id,
                        date=attendance_date,
                        status=status,
                        week_number=week_number,
                        year=year,
                        notes=notes
                    )
                    db.session.add(attendance)
        db.session.commit()
        log_activity(current_user, 'create', 'AttendanceRecord', None, f"سجل حضور/غياب جماعي بتاريخ {attendance_date}")
        flash('تم تسجيل الحضور/الغياب الجماعي بنجاح!', 'success')
        return redirect(url_for('bulk_attendance'))
    settings = CompanySettings.query.first()
    return render_template('salary/bulk_attendance.html', employees=employees, settings=settings)

# --- تسجيل الساعات والأيام الإضافية الجماعية ---
@app.route('/salary/overtime/bulk', methods=['GET', 'POST'])
@login_required
def bulk_overtime():
    if not current_user.can_edit():
        flash('ليس لديك صلاحية تسجيل الساعات الإضافية.', 'danger')
        return redirect(url_for('index'))
    employees = Employee.query.filter_by(status='active').all()
    if request.method == 'POST':
        date_str = request.form['date']
        overtime_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        week_number = overtime_date.isocalendar()[1]
        year = overtime_date.year
        for employee in employees:
            overtime_type = request.form.get(f'overtime_type_{employee.id}')
            quantity = request.form.get(f'quantity_{employee.id}')
            notes = request.form.get(f'notes_{employee.id}')
            if overtime_type and quantity and float(quantity) > 0:
                overtime = OvertimeRecord(
                    employee_id=employee.id,
                    date=overtime_date,
                    overtime_type=overtime_type,
                    quantity=float(quantity),
                    week_number=week_number,
                    year=year,
                    notes=notes
                )
                db.session.add(overtime)
        db.session.commit()
        log_activity(current_user, 'create', 'OvertimeRecord', None, f"سجل ساعات إضافية جماعية بتاريخ {overtime_date}")
        flash('تم تسجيل الساعات الإضافية الجماعية بنجاح!', 'success')
        return redirect(url_for('bulk_overtime'))
    settings = CompanySettings.query.first()
    return render_template('salary/bulk_overtime.html', employees=employees, settings=settings)

# --- تسجيل السلف الجماعية ---
@app.route('/salary/advance/bulk', methods=['GET', 'POST'])
@login_required
def bulk_advance():
    if not current_user.can_edit():
        flash('ليس لديك صلاحية تسجيل السلف.', 'danger')
        return redirect(url_for('index'))
    employees = Employee.query.filter_by(status='active').all()
    if request.method == 'POST':
        payment_date_str = request.form['payment_date']
        payment_date = datetime.strptime(payment_date_str, '%Y-%m-%d').date()
        for employee in employees:
            amount = request.form.get(f'amount_{employee.id}')
            reason = request.form.get(f'reason_{employee.id}')
            if amount and float(amount) > 0:
                advance = AdvancePayment(
                    employee_id=employee.id,
                    amount=float(amount),
                    payment_date=payment_date,
                    reason=reason,
                    is_paid=False
                )
                db.session.add(advance)
        db.session.commit()
        log_activity(current_user, 'create', 'AdvancePayment', None, f"سجل سلف جماعية بتاريخ {payment_date}")
        flash('تم تسجيل السلف الجماعية بنجاح!', 'success')
        return redirect(url_for('bulk_advance'))
    settings = CompanySettings.query.first()
    return render_template('salary/bulk_advance.html', employees=employees, settings=settings)

# --- حساب الرواتب الأسبوعية ---
@app.route('/salary/payroll/calculate', methods=['GET', 'POST'])
@login_required
def calculate_payroll():
    if not current_user.is_admin():
        flash('ليس لديك صلاحية حساب الرواتب.', 'danger')
        return redirect(url_for('index'))
    salary_settings = SalarySettings.query.first()
    if not salary_settings:
        flash('يرجى تعيين إعدادات الرواتب أولاً.', 'warning')
        return redirect(url_for('salary_settings'))
    if request.method == 'POST':
        week_number = int(request.form['week_number'])
        year = int(request.form['year'])
        employee_id = request.form.get('employee_id')
        if employee_id:
            employees = [Employee.query.get_or_404(int(employee_id))]
        else:
            employees = Employee.query.filter_by(status='active').all()
        for employee in employees:
            attendances = AttendanceRecord.query.filter_by(
                employee_id=employee.id,
                week_number=week_number,
                year=year
            ).all()
            overtime_records = OvertimeRecord.query.filter_by(
                employee_id=employee.id,
                week_number=week_number,
                year=year
            ).all()
            advances = AdvancePayment.query.filter_by(
                employee_id=employee.id,
                is_paid=False
            ).all()
            total_days = 6
            present_days = len([a for a in attendances if a.status == 'present'])
            absent_days = len([a for a in attendances if a.status == 'absent'])
            half_days = len([a for a in attendances if a.status == 'half_day'])
            overtime_days = sum([r.quantity for r in overtime_records if r.overtime_type == 'daily'])
            overtime_hours = sum([r.quantity for r in overtime_records if r.overtime_type == 'hourly'])
            salary_info = EmployeeSalary.query.filter_by(employee_id=employee.id).first()
            if not salary_info:
                flash(f'لم يتم تعيين معلومات الراتب للموظف: {employee.full_name}', 'warning')
                continue
            if salary_info.daily_wage:
                basic_salary = present_days * salary_info.daily_wage
                if half_days > 0:
                    basic_salary += half_days * (salary_info.daily_wage / 2)
            elif salary_info.base_salary:
                daily_rate = salary_info.base_salary / 26
                basic_salary = present_days * daily_rate
                if half_days > 0:
                    basic_salary += half_days * (daily_rate / 2)
            else:
                basic_salary = 0
            overtime_amount = (overtime_days * salary_settings.overtime_daily_rate) + (overtime_hours * salary_settings.overtime_hourly_rate)
            deductions = absent_days * (salary_info.daily_wage if salary_info.daily_wage else salary_info.base_salary / 26)
            advances_deduction = sum([a.amount for a in advances])
            net_salary = basic_salary + overtime_amount - deductions - advances_deduction
            existing = PayrollRecord.query.filter_by(
                employee_id=employee.id,
                week_number=week_number,
                year=year
            ).first()
            if existing:
                existing.present_days = present_days
                existing.absent_days = absent_days
                existing.half_days = half_days
                existing.overtime_days = overtime_days
                existing.overtime_hours = overtime_hours
                existing.basic_salary = basic_salary
                existing.overtime_amount = overtime_amount
                existing.deductions = deductions
                existing.advances_deduction = advances_deduction
                existing.net_salary = net_salary
            else:
                payroll = PayrollRecord(
                    employee_id=employee.id,
                    week_number=week_number,
                    year=year,
                    present_days=present_days,
                    absent_days=absent_days,
                    half_days=half_days,
                    overtime_days=overtime_days,
                    overtime_hours=overtime_hours,
                    basic_salary=basic_salary,
                    overtime_amount=overtime_amount,
                    deductions=deductions,
                    advances_deduction=advances_deduction,
                    net_salary=net_salary,
                    paid=False
                )
                db.session.add(payroll)
        db.session.commit()
        flash('تم حساب الرواتب بنجاح!', 'success')
        return redirect(url_for('payroll_list', week_number=week_number, year=year))
    weeks = db.session.query(AttendanceRecord.week_number, AttendanceRecord.year).distinct().order_by(AttendanceRecord.year.desc(), AttendanceRecord.week_number.desc()).all()
    employees = Employee.query.filter_by(status='active').all()
    settings = CompanySettings.query.first()
    return render_template('salary/payroll_calculate.html', weeks=weeks, employees=employees, settings=settings)

# --- قائمة الرواتب ---
@app.route('/salary/payroll')
@login_required
def payroll_list():
    week_number = request.args.get('week_number', type=int)
    year = request.args.get('year', type=int)
    if week_number and year:
        payroll_records = PayrollRecord.query.filter_by(week_number=week_number, year=year).all()
    else:
        payroll_records = PayrollRecord.query.order_by(PayrollRecord.year.desc(), PayrollRecord.week_number.desc()).all()
    settings = CompanySettings.query.first()
    return render_template('salary/payroll_list.html', payroll_records=payroll_records, settings=settings)

# --- تفاصيل راتب موظف ---
@app.route('/salary/payroll/<int:record_id>')
@login_required
def payroll_detail(record_id):
    record = PayrollRecord.query.get_or_404(record_id)
    settings = CompanySettings.query.first()
    return render_template('salary/payroll_detail.html', record=record, settings=settings)

# --- دفع الراتب ---
@app.route('/salary/payroll/<int:record_id>/pay', methods=['POST'])
@login_required
def pay_payroll(record_id):
    if not current_user.is_admin():
        flash('ليس لديك صلاحية دفع الرواتب.', 'danger')
        return redirect(url_for('payroll_list'))
    record = PayrollRecord.query.get_or_404(record_id)
    record.paid = True
    record.paid_date = datetime.now().date()
    db.session.commit()
    advances = AdvancePayment.query.filter_by(
        employee_id=record.employee_id,
        is_paid=False
    ).all()
    for advance in advances:
        advance.is_paid = True
        advance.paid_date = datetime.now().date()
    db.session.commit()
    flash('تم دفع الراتب بنجاح!', 'success')
    return redirect(url_for('payroll_detail', record_id=record.id))

# --- تصدير الرواتب إلى Excel ---
@app.route('/salary/payroll/export')
@login_required
def export_payroll():
    week_number = request.args.get('week_number', type=int)
    year = request.args.get('year', type=int)
    if week_number and year:
        records = PayrollRecord.query.filter_by(week_number=week_number, year=year).all()
    else:
        records = PayrollRecord.query.all()
    data = [{
        'اسم الموظف': record.employee.full_name,
        'الأسبوع': f"{record.week_number}-{record.year}",
        'أيام الحضور': record.present_days,
        'أيام الغياب': record.absent_days,
        'نصف أيام': record.half_days,
        'أيام إضافية': record.overtime_days,
        'ساعات إضافية': record.overtime_hours,
        'الراتب الأساسي': record.basic_salary,
        'بدل الساعات الإضافية': record.overtime_amount,
        'خصم الغياب': record.deductions,
        'حسم السلف': record.advances_deduction,
        'صافي الراتب': record.net_salary,
        'مدفوع': 'نعم' if record.paid else 'لا',
        'تاريخ الدفع': record.paid_date.strftime('%Y-%m-%d') if record.paid_date else '-'
    } for record in records]
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='رواتب الموظفين')
        workbook = writer.book
        worksheet = writer.sheets['رواتب الموظفين']
        settings = CompanySettings.query.first()
        company_name = settings.company_name if settings else "شركة الأرشيف"
        worksheet['A1'] = company_name
        worksheet['A1'].font = Font(size=16, bold=True, color="0070C0")
        worksheet.merge_cells('A1:O1')
        worksheet['A1'].alignment = Alignment(horizontal="center")
    output.seek(0)
    filename = f'payroll_export_{week_number or "all"}_{year or "all"}.xlsx'
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
# --- إدارة المستودعات ---
@app.route('/warehouses')
@login_required
def warehouse_list():
    warehouses = Warehouse.query.all()
    settings = CompanySettings.query.first()
    return render_template('warehouse/list.html', warehouses=warehouses, settings=settings)

@app.route('/warehouses/add', methods=['GET', 'POST'])
@login_required
def add_warehouse():
    if not current_user.is_admin():
        flash('ليس لديك صلاحية لإضافة مستودعات.', 'danger')
        return redirect(url_for('warehouse_list'))
    if request.method == 'POST':
        name = request.form['name']
        location = request.form.get('location')
        is_active = bool(request.form.get('is_active'))
        if Warehouse.query.filter_by(name=name).first():
            flash('اسم المستودع موجود مسبقاً.', 'warning')
            return render_template('warehouse/add.html')
        warehouse = Warehouse(name=name, location=location, is_active=is_active)
        db.session.add(warehouse)
        db.session.commit()
        log_activity(current_user, 'create', 'Warehouse', warehouse.id, f"أضاف مستودع: {name}")
        flash('تم إضافة المستودع بنجاح!', 'success')
        return redirect(url_for('warehouse_list'))
    settings = CompanySettings.query.first()
    return render_template('warehouse/add.html', settings=settings)

@app.route('/warehouses/<int:warehouse_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_warehouse(warehouse_id):
    if not current_user.is_admin():
        flash('ليس لديك صلاحية تعديل المستودعات.', 'danger')
        return redirect(url_for('warehouse_list'))
    warehouse = Warehouse.query.get_or_404(warehouse_id)
    if request.method == 'POST':
        warehouse.name = request.form['name']
        warehouse.location = request.form.get('location')
        warehouse.is_active = bool(request.form.get('is_active'))
        db.session.commit()
        log_activity(current_user, 'update', 'Warehouse', warehouse.id, f"عدل مستودع: {warehouse.name}")
        flash('تم تعديل المستودع بنجاح!', 'success')
        return redirect(url_for('warehouse_list'))
    settings = CompanySettings.query.first()
    return render_template('warehouse/edit.html', warehouse=warehouse, settings=settings)

@app.route('/warehouses/<int:warehouse_id>/delete', methods=['POST'])
@login_required
def delete_warehouse(warehouse_id):
    if not current_user.is_admin():
        flash('ليس لديك صلاحية حذف المستودعات.', 'danger')
        return redirect(url_for('warehouse_list'))
    warehouse = Warehouse.query.get_or_404(warehouse_id)
    if StockItem.query.filter_by(warehouse_id=warehouse_id).first():
        flash('لا يمكن حذف مستودع يحتوي على مواد.', 'warning')
        return redirect(url_for('warehouse_list'))
    name = warehouse.name
    db.session.delete(warehouse)
    db.session.commit()
    log_activity(current_user, 'delete', 'Warehouse', warehouse_id, f"حذف مستودع: {name}")
    flash('تم حذف المستودع بنجاح!', 'success')
    return redirect(url_for('warehouse_list'))

# --- إدارة المواد ---
@app.route('/materials')
@login_required
def material_list():
    materials = Material.query.all()
    settings = CompanySettings.query.first()
    return render_template('material/list.html', materials=materials, settings=settings)

@app.route('/materials/add', methods=['GET', 'POST'])
@login_required
def add_material():
    if not current_user.is_admin():
        flash('ليس لديك صلاحية لإضافة مواد.', 'danger')
        return redirect(url_for('material_list'))
    if request.method == 'POST':
        name = request.form['name']
        unit = request.form['unit']
        min_stock_level = request.form.get('min_stock_level', 0)
        category = request.form.get('category')
        notes = request.form.get('notes')
        if Material.query.filter_by(name=name).first():
            flash('اسم المادة موجود مسبقاً.', 'warning')
            return render_template('material/add.html')
        material = Material(
            name=name,
            unit=unit,
            min_stock_level=float(min_stock_level),
            category=category,
            notes=notes
        )
        db.session.add(material)
        db.session.commit()
        log_activity(current_user, 'create', 'Material', material.id, f"أضاف مادة: {name}")
        flash('تم إضافة المادة بنجاح!', 'success')
        return redirect(url_for('material_list'))
    settings = CompanySettings.query.first()
    return render_template('material/add.html', settings=settings)

@app.route('/materials/<int:material_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_material(material_id):
    if not current_user.is_admin():
        flash('ليس لديك صلاحية تعديل المواد.', 'danger')
        return redirect(url_for('material_list'))
    material = Material.query.get_or_404(material_id)
    if request.method == 'POST':
        material.name = request.form['name']
        material.unit = request.form['unit']
        material.min_stock_level = float(request.form.get('min_stock_level', 0))
        material.category = request.form.get('category')
        material.notes = request.form.get('notes')
        db.session.commit()
        log_activity(current_user, 'update', 'Material', material.id, f"عدل مادة: {material.name}")
        flash('تم تعديل المادة بنجاح!', 'success')
        return redirect(url_for('material_list'))
    settings = CompanySettings.query.first()
    return render_template('material/edit.html', material=material, settings=settings)

@app.route('/materials/<int:material_id>/delete', methods=['POST'])
@login_required
def delete_material(material_id):
    if not current_user.is_admin():
        flash('ليس لديك صلاحية حذف المواد.', 'danger')
        return redirect(url_for('material_list'))
    material = Material.query.get_or_404(material_id)
    if StockItem.query.filter_by(material_id=material_id).first():
        flash('لا يمكن حذف مادة لها رصيد في المخازن.', 'warning')
        return redirect(url_for('material_list'))
    name = material.name
    db.session.delete(material)
    db.session.commit()
    log_activity(current_user, 'delete', 'Material', material_id, f"حذف مادة: {name}")
    flash('تم حذف المادة بنجاح!', 'success')
    return redirect(url_for('material_list'))

# --- حركة المخزون (إضافة / صرف) ---
@app.route('/inventory/transaction', methods=['GET', 'POST'])
@login_required
def add_stock_transaction():
    if not current_user.can_edit():
        flash('ليس لديك صلاحية إدارة المخزون.', 'danger')
        return redirect(url_for('index'))
    warehouses = Warehouse.query.filter_by(is_active=True).all()
    materials = Material.query.all()
    if request.method == 'POST':
        warehouse_id = int(request.form['warehouse_id'])
        material_id = int(request.form['material_id'])
        transaction_type = request.form['transaction_type']
        quantity = float(request.form['quantity'])
        reference = request.form.get('reference')
        notes = request.form.get('notes')
        # الحصول على رصيد المادة في هذا المستودع
        stock_item = StockItem.query.filter_by(warehouse_id=warehouse_id, material_id=material_id).first()
        if not stock_item:
            stock_item = StockItem(warehouse_id=warehouse_id, material_id=material_id, quantity=0)
            db.session.add(stock_item)
        # تعديل الرصيد
        if transaction_type == 'in':
            stock_item.quantity += quantity
        elif transaction_type == 'out':
            if stock_item.quantity < quantity:
                flash('الكمية المطلوبة غير متوفرة في المخزن!', 'danger')
                return redirect(url_for('add_stock_transaction'))
            stock_item.quantity -= quantity
        stock_item.last_updated = datetime.utcnow()
        # تسجيل الحركة
        transaction = StockTransaction(
            warehouse_id=warehouse_id,
            material_id=material_id,
            transaction_type=transaction_type,
            quantity=quantity,
            balance_after=stock_item.quantity,
            reference=reference,
            notes=notes,
            created_by_id=current_user.id
        )
        db.session.add(transaction)
        db.session.commit()
        action = "إضافة" if transaction_type == 'in' else "صرف"
        log_activity(current_user, 'create', 'StockTransaction', transaction.id, f"{action} {quantity} {stock_item.material.unit} من {stock_item.material.name} في {stock_item.warehouse.name}")
        flash(f'تم {action} الكمية بنجاح!', 'success')
        return redirect(url_for('stock_balance'))
    settings = CompanySettings.query.first()
    return render_template('inventory/transaction.html', warehouses=warehouses, materials=materials, settings=settings)

# --- رصيد المخزون ---
@app.route('/inventory/balance')
@login_required
def stock_balance():
    # جلب جميع أرصدة المخزون مع أسماء المستودعات والمواد
    balances = db.session.query(
        StockItem.id,
        Warehouse.name.label('warehouse_name'),
        Material.name.label('material_name'),
        Material.unit,
        Material.min_stock_level,
        StockItem.quantity
    ).join(Warehouse, StockItem.warehouse_id == Warehouse.id)\
     .join(Material, StockItem.material_id == Material.id)\
     .all()
    # تصنيف المواد التي تحت الحد الأدنى
    low_stock_items = [b for b in balances if b.quantity <= b.min_stock_level]
    settings = CompanySettings.query.first()
    return render_template('inventory/balance.html', balances=balances, low_stock_items=low_stock_items, settings=settings)

# --- حركة مادة معينة ---
@app.route('/inventory/material/<int:material_id>')
@login_required
def material_history(material_id):
    material = Material.query.get_or_404(material_id)
    transactions = StockTransaction.query.filter_by(material_id=material_id)\
        .order_by(StockTransaction.created_at.desc()).all()
    settings = CompanySettings.query.first()
    return render_template('inventory/material_history.html', material=material, transactions=transactions, settings=settings)

# --- تصدير رصيد المخزون إلى Excel ---
@app.route('/inventory/export/balance')
@login_required
def export_stock_balance():
    balances = db.session.query(
        Warehouse.name.label('warehouse_name'),
        Material.name.label('material_name'),
        Material.unit,
        StockItem.quantity
    ).join(Warehouse, StockItem.warehouse_id == Warehouse.id)\
     .join(Material, StockItem.material_id == Material.id)\
     .all()
    data = [{
        'المستودع': b.warehouse_name,
        'المادة': b.material_name,
        'الوحدة': b.unit,
        'الكمية': b.quantity
    } for b in balances]
    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='رصيد المخزون')
        workbook = writer.book
        worksheet = writer.sheets['رصيد المخزون']
        settings = CompanySettings.query.first()
        company_name = settings.company_name if settings else "شركة الأرشيف"
        worksheet['A1'] = company_name
        worksheet['A1'].font = Font(size=16, bold=True, color="0070C0")
        worksheet.merge_cells('A1:D1')
        worksheet['A1'].alignment = Alignment(horizontal="center")
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name='stock_balance.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
# --- Context Processor ---
@app.context_processor
def inject_current_year():
    return {'current_year': datetime.now().year}

import atexit
atexit.register(lambda: scheduler.shutdown())
# --- نموذج المستودع ---
class Warehouse(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    location = db.Column(db.String(200))
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    # العلاقات
    stock_items = db.relationship('StockItem', backref='warehouse', lazy=True, cascade="all, delete-orphan")
    transactions = db.relationship('StockTransaction', backref='warehouse', lazy=True, cascade="all, delete-orphan")

    def __repr__(self):
        return f'<Warehouse {self.name}>'

# --- نموذج المادة (المنتج) ---
class Material(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    unit = db.Column(db.String(20), nullable=False)  # وحدة القياس: طن، متر مكعب، لتر، قطعة...
    min_stock_level = db.Column(db.Float, default=0)  # الحد الأدنى للتنبيه
    category = db.Column(db.String(50))  # مواد خام، قطع غيار، وقود...
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    # العلاقات
    stock_items = db.relationship('StockItem', backref='material', lazy=True, cascade="all, delete-orphan")
    transactions = db.relationship('StockTransaction', backref='material', lazy=True, cascade="all, delete-orphan")

    def __repr__(self):
        return f'<Material {self.name}>'

# --- نموذج رصيد المخزون (رصيد المادة في مستودع معين) ---
class StockItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    warehouse_id = db.Column(db.Integer, db.ForeignKey('warehouse.id'), nullable=False)
    material_id = db.Column(db.Integer, db.ForeignKey('material.id'), nullable=False)
    quantity = db.Column(db.Float, default=0)  # الكمية المتوفرة
    last_updated = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # ضمان فريد: مادة واحدة في مستودع واحد
    __table_args__ = (db.UniqueConstraint('warehouse_id', 'material_id', name='uq_warehouse_material'),)

    def __repr__(self):
        return f'<StockItem {self.material.name} in {self.warehouse.name}: {self.quantity} {self.material.unit}>'

# --- نموذج حركة المخزون (سجل العمليات) ---
class StockTransaction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    warehouse_id = db.Column(db.Integer, db.ForeignKey('warehouse.id'), nullable=False)
    material_id = db.Column(db.Integer, db.ForeignKey('material.id'), nullable=False)
    transaction_type = db.Column(db.String(20), nullable=False)  # 'in' (دخول) أو 'out' (صرف)
    quantity = db.Column(db.Float, nullable=False)
    balance_after = db.Column(db.Float, nullable=False)  # الرصيد بعد هذه العملية
    reference = db.Column(db.String(100))  # رقم فاتورة — أمر شغل — إلخ
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)

    # العلاقات
    created_by = db.relationship('User', backref='stock_transactions')

    def __repr__(self):
        return f'<StockTransaction {self.transaction_type} {self.quantity} of {self.material.name}>'
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)